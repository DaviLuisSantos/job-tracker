from datetime import datetime
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, Rule
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.filters import AutoFilter
from rich.console import Console

from db.database import fetch_all_jobs, fetch_scrape_runs

console = Console()

OUTPUT_DIR = Path(__file__).parent.parent / "data"

# ── Paleta de cores ───────────────────────────────────────────────────────────
C_HEADER_BG   = "1F3864"   # azul escuro
C_HEADER_FG   = "FFFFFF"
C_SUBHEADER   = "2E75B6"
C_ACCENT      = "4472C4"
C_GREEN_BG    = "E2EFDA"
C_YELLOW_BG   = "FFEB9C"
C_RED_BG      = "FFC7CE"
C_ALT_ROW     = "F2F2F2"
C_BORDER      = "BFBFBF"

PROFILE_COLORS = {
    "tech":         "2E75B6",
    "trades":       "375623",
    "construction": "7B3F00",
    "fifo":         "7030A0",
    "default":      "595959",
}

# ── Helpers de estilo ─────────────────────────────────────────────────────────

def _thin_border():
    s = Side(style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)


def _header_style(hex_bg: str = C_HEADER_BG):
    return {
        "font":      Font(name="Arial", bold=True, color=C_HEADER_FG, size=10),
        "fill":      PatternFill("solid", fgColor=hex_bg),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "border":    _thin_border(),
    }


def _apply(cell, **kwargs):
    for attr, val in kwargs.items():
        setattr(cell, attr, val)


def _score_fill(score: int) -> PatternFill:
    if score >= 80:
        return PatternFill("solid", fgColor="C6EFCE")
    if score >= 70:
        return PatternFill("solid", fgColor="E2EFDA")
    if score >= 50:
        return PatternFill("solid", fgColor="FFEB9C")
    return PatternFill("solid", fgColor="FCE4D6")


def _score_font(score: int) -> Font:
    color = "375623" if score >= 70 else ("7B3F00" if score >= 50 else "9C0006")
    return Font(name="Arial", bold=True, color=color, size=10)


def _autofit(ws, min_w=8, max_w=50):
    for col in ws.columns:
        length = max(
            (len(str(cell.value or "")) for cell in col),
            default=min_w,
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(
            max(length + 2, min_w), max_w
        )


# ── Aba: Resumo ───────────────────────────────────────────────────────────────

def _build_summary(wb: Workbook, all_jobs: list[dict], runs: list[dict]):
    ws = wb.active
    ws.title = "Resumo"
    ws.sheet_view.showGridLines = False

    # Título principal
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = f"Job Tracker — Relatório gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    _apply(title_cell,
           font=Font(name="Arial", bold=True, size=14, color=C_HEADER_FG),
           fill=PatternFill("solid", fgColor=C_HEADER_BG),
           alignment=Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[1].height = 28

    # ── Stats gerais ──────────────────────────────────────────
    df = pd.DataFrame(all_jobs) if all_jobs else pd.DataFrame()

    stats_headers = ["Métrica", "Valor"]
    stats_data = []
    if not df.empty:
        stats_data = [
            ("Total de vagas",       len(df)),
            ("Novas (não lidas)",     int((df["status"] == "new").sum())),
            ("Alto fit (≥ 70)",       int((df["fit_score"] >= 70).sum())),
            ("Candidaturas enviadas", int(df["applied"].sum())),
            ("Score médio",           round(float(df["fit_score"].mean()), 1)),
            ("Score máximo",          int(df["fit_score"].max())),
        ]

    ws.append([])  # linha 2 em branco
    ws.row_dimensions[2].height = 6

    # Cabeçalho stats
    ws.append(stats_headers)
    header_row = ws.max_row
    for col, h in enumerate(stats_headers, 1):
        cell = ws.cell(header_row, col)
        _apply(cell, **_header_style())

    for metric, value in stats_data:
        ws.append([metric, value])
        row = ws.max_row
        ws.cell(row, 1).font = Font(name="Arial", size=10)
        ws.cell(row, 1).border = _thin_border()
        ws.cell(row, 2).font = Font(name="Arial", bold=True, size=10)
        ws.cell(row, 2).alignment = Alignment(horizontal="center")
        ws.cell(row, 2).border = _thin_border()
        if row % 2 == 0:
            for c in [1, 2]:
                ws.cell(row, c).fill = PatternFill("solid", fgColor=C_ALT_ROW)

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 16

    # ── Por perfil ────────────────────────────────────────────
    if not df.empty and "profile" in df.columns:
        ws.append([])
        ws.append([])
        ws.row_dimensions[ws.max_row].height = 6

        profile_headers = ["Perfil", "Total", "Novas", "Alto fit (≥70)", "Score médio", "Score máx"]
        ws.append(profile_headers)
        ph_row = ws.max_row
        for col, h in enumerate(profile_headers, 1):
            cell = ws.cell(ph_row, col)
            _apply(cell, **_header_style())

        for profile, grp in df.groupby("profile"):
            color = PROFILE_COLORS.get(profile, "595959")
            row_data = [
                profile,
                len(grp),
                int((grp["status"] == "new").sum()),
                int((grp["fit_score"] >= 70).sum()),
                round(float(grp["fit_score"].mean()), 1),
                int(grp["fit_score"].max()),
            ]
            ws.append(row_data)
            r = ws.max_row
            ws.cell(r, 1).font = Font(name="Arial", bold=True, color=color, size=10)
            for c in range(1, len(profile_headers) + 1):
                ws.cell(r, c).border = _thin_border()
                ws.cell(r, c).alignment = Alignment(horizontal="center")
            ws.cell(r, 1).alignment = Alignment(horizontal="left")
            if r % 2 == 0:
                for c in range(1, len(profile_headers) + 1):
                    ws.cell(r, c).fill = PatternFill("solid", fgColor=C_ALT_ROW)

        for col_idx, w in enumerate([18, 10, 10, 14, 14, 12], 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = w

    # ── Últimas execuções ─────────────────────────────────────
    if runs:
        ws.append([])
        ws.append([])
        ws.row_dimensions[ws.max_row].height = 6

        run_headers = ["Data/Hora", "Perfil", "Keyword", "Novas", "Total encontrado"]
        ws.append(run_headers)
        rh_row = ws.max_row
        for col, h in enumerate(run_headers, 1):
            _apply(ws.cell(rh_row, col), **_header_style(C_SUBHEADER))

        for run in runs[:20]:
            ws.append([
                run.get("ran_at", ""),
                run.get("profile", "—"),
                run.get("keyword", ""),
                run.get("new_jobs", 0),
                run.get("found", 0),
            ])
            r = ws.max_row
            for c in range(1, 6):
                ws.cell(r, c).font = Font(name="Arial", size=9)
                ws.cell(r, c).border = _thin_border()
            if r % 2 == 0:
                for c in range(1, 6):
                    ws.cell(r, c).fill = PatternFill("solid", fgColor=C_ALT_ROW)


# ── Aba: por perfil ───────────────────────────────────────────────────────────

JOB_COLUMNS = [
    ("fit_score",        "Score",          7),
    ("title",            "Título",         35),
    ("company",          "Empresa",        22),
    ("location",         "Localização",    20),
    ("salary",           "Salário",        22),
    ("salary_annual_min","Sal. Anual Mín", 16),
    ("salary_annual_max","Sal. Anual Máx", 16),
    ("work_type",        "Tipo",           14),
    ("status",           "Status",         12),
    ("listed_at",        "Publicado em",   18),
    ("url",              "Link",           12),
]


def _build_profile_sheet(wb: Workbook, profile_name: str, jobs: list[dict]):
    color = PROFILE_COLORS.get(profile_name, "595959")
    ws = wb.create_sheet(title=profile_name.capitalize())
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    # Cabeçalho
    headers = [col[1] for col in JOB_COLUMNS]
    ws.append(headers)
    for col_idx, (_, label, width) in enumerate(JOB_COLUMNS, 1):
        cell = ws.cell(1, col_idx)
        _apply(cell, **_header_style(color))
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 20

    # Dados
    df = pd.DataFrame(jobs).sort_values("fit_score", ascending=False)
    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        score = int(row.get("fit_score") or 0)
        alt   = row_idx % 2 == 0

        for col_idx, (field, _, _) in enumerate(JOB_COLUMNS, 1):
            cell = ws.cell(row_idx, col_idx)
            val  = row.get(field)

            if field == "fit_score":
                cell.value = score
                cell.fill  = _score_fill(score)
                cell.font  = _score_font(score)
                cell.alignment = Alignment(horizontal="center")
            elif field in ("salary_annual_min", "salary_annual_max"):
                if val is not None and not pd.isna(val):
                    cell.value        = int(val)
                    cell.number_format = '$#,##0'
                    cell.alignment    = Alignment(horizontal="right")
                else:
                    cell.value = ""
                cell.font = Font(name="Arial", size=9)
                if alt:
                    cell.fill = PatternFill("solid", fgColor=C_ALT_ROW)
            elif field == "url" and val:
                cell.value     = "Abrir vaga"
                cell.hyperlink = str(val)
                cell.font      = Font(name="Arial", color="0563C1", underline="single", size=9)
                cell.alignment = Alignment(horizontal="center")
            elif field == "listed_at" and val:
                try:
                    cell.value = str(val)[:10]
                except Exception:
                    cell.value = str(val) if val else ""
                cell.alignment = Alignment(horizontal="center")
                cell.font = Font(name="Arial", size=9)
            else:
                cell.value = str(val) if val else ""
                cell.font  = Font(name="Arial", size=9)
                if alt and field != "fit_score":
                    cell.fill = PatternFill("solid", fgColor=C_ALT_ROW)

            cell.border = _thin_border()

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(JOB_COLUMNS))}1"

    # Legenda de score
    legend_col = len(JOB_COLUMNS) + 2
    ws.cell(1, legend_col).value = "Legenda Score"
    _apply(ws.cell(1, legend_col),
           font=Font(name="Arial", bold=True, size=9),
           alignment=Alignment(horizontal="center"))
    ws.column_dimensions[get_column_letter(legend_col)].width = 16

    for r, (label, fill_color) in enumerate([
        ("≥ 80 — Excelente", "C6EFCE"),
        ("70–79 — Bom",      "E2EFDA"),
        ("50–69 — Regular",  "FFEB9C"),
        ("< 50 — Baixo",     "FCE4D6"),
    ], 2):
        cell = ws.cell(r, legend_col)
        cell.value = label
        cell.fill  = PatternFill("solid", fgColor=fill_color)
        cell.font  = Font(name="Arial", size=9)
        cell.border = _thin_border()
        cell.alignment = Alignment(horizontal="center")


# ── Aba: todas as vagas ───────────────────────────────────────────────────────

def _build_all_sheet(wb: Workbook, all_jobs: list[dict]):
    ws = wb.create_sheet(title="Todas as vagas")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    cols = [("profile", "Perfil", 14)] + JOB_COLUMNS
    ws.append([c[1] for c in cols])
    for col_idx, (_, label, width) in enumerate(cols, 1):
        _apply(ws.cell(1, col_idx), **_header_style())
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 20

    df = pd.DataFrame(all_jobs).sort_values(["profile", "fit_score"], ascending=[True, False])

    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        score = int(row.get("fit_score") or 0)
        alt   = row_idx % 2 == 0
        profile = str(row.get("profile") or "")

        for col_idx, (field, _, _) in enumerate(cols, 1):
            cell = ws.cell(row_idx, col_idx)
            val  = row.get(field)

            if field == "profile":
                cell.value = profile
                color = PROFILE_COLORS.get(profile, "595959")
                cell.font = Font(name="Arial", bold=True, color=color, size=9)
            elif field == "fit_score":
                cell.value = score
                cell.fill  = _score_fill(score)
                cell.font  = _score_font(score)
                cell.alignment = Alignment(horizontal="center")
            elif field in ("salary_annual_min", "salary_annual_max"):
                if val is not None and not pd.isna(val):
                    cell.value         = int(val)
                    cell.number_format = '$#,##0'
                    cell.alignment     = Alignment(horizontal="right")
                else:
                    cell.value = ""
                cell.font = Font(name="Arial", size=9)
                if alt:
                    cell.fill = PatternFill("solid", fgColor=C_ALT_ROW)
            elif field == "url" and val:
                cell.value     = "Abrir vaga"
                cell.hyperlink = str(val)
                cell.font      = Font(name="Arial", color="0563C1", underline="single", size=9)
                cell.alignment = Alignment(horizontal="center")
            elif field == "listed_at" and val:
                cell.value = str(val)[:10]
                cell.font  = Font(name="Arial", size=9)
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.value = str(val) if val else ""
                cell.font  = Font(name="Arial", size=9)
                if alt and field not in ("fit_score",):
                    cell.fill = PatternFill("solid", fgColor=C_ALT_ROW)

            cell.border = _thin_border()

    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"


# ── Entry point ───────────────────────────────────────────────────────────────

def export_excel(profile: str | None = None) -> Path:
    all_jobs = fetch_all_jobs(profile)
    runs     = fetch_scrape_runs(profile, limit=20)

    if not all_jobs:
        console.print("[yellow]Nenhuma vaga para exportar.[/yellow]")
        return None

    wb = Workbook()

    _build_summary(wb, all_jobs, runs)

    # Uma aba por perfil
    df = pd.DataFrame(all_jobs)
    profiles_present = df["profile"].unique() if "profile" in df.columns else []

    for pname in sorted(profiles_present):
        jobs_for_profile = [j for j in all_jobs if j.get("profile") == pname]
        _build_profile_sheet(wb, pname, jobs_for_profile)

    # Aba consolidada (só quando há mais de um perfil)
    if len(profiles_present) > 1:
        _build_all_sheet(wb, all_jobs)

    # Salva
    OUTPUT_DIR.mkdir(exist_ok=True)
    stamp    = datetime.now().strftime("%Y-%m-%d_%H-%M")
    suffix   = f"_{profile}" if profile else ""
    out_path = OUTPUT_DIR / f"job_report{suffix}_{stamp}.xlsx"
    wb.save(out_path)

    console.print(f"[bold green]Relatorio salvo em: {out_path}[/bold green]")
    return out_path
